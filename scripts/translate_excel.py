"""
Excel翻译脚本 - 百度翻译API版本
支持批量翻译，一次性处理多行
"""

import openpyxl
from deep_translator.baidu import BaiduTranslator
import sys
import time
import os
import gc


def translate_batch(translator, texts):
    """批量翻译多行文本"""
    if not texts:
        return []

    # 用换行符连接所有文本
    combined = "\n".join(texts)
    try:
        result = translator.translate(combined)
        # 返回的结果也是用换行符分隔的
        return result.split("\n") if result else []
    except Exception as e:
        print(f"批量翻译错误: {e}")
        return None


def translate_excel(input_file, output_file, appid, appkey, batch_size=100):
    """
    翻译Excel文件的C列和D列
    批量翻译，每batch_size行处理一次
    """
    # 检查进度
    start_row = 2

    if os.path.exists(output_file):
        try:
            temp_wb = openpyxl.load_workbook(output_file)
            temp_ws = temp_wb.active
            for i in range(2, temp_ws.max_row + 1):
                c_val = temp_ws.cell(i, 3).value
                if c_val and isinstance(c_val, str):
                    has_cn = any('\u4e00' <= c <= '\u9fff' for c in c_val)
                    if has_cn:
                        start_row = i
                        break
            else:
                start_row = temp_ws.max_row + 1
            temp_wb.close()
            print(f"从第 {start_row} 行继续...")
        except:
            start_row = 2

    if not os.path.exists(output_file):
        wb = openpyxl.load_workbook(input_file)
        wb.save(output_file)
        wb.close()
        print(f"创建新文件: {output_file}")

    # 读取源文件
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    total_rows = ws.max_row - 1

    # 百度翻译器 (语言代码: zh 而不是 zh-CN)
    translator = BaiduTranslator(source='zh', target='en', appid=appid, appkey=appkey)

    print(f"总共 {total_rows} 行待翻译")
    print(f"批量大小: {batch_size}行/次")

    count = 0
    row_idx = start_row

    # 翻译表头
    output_wb = openpyxl.load_workbook(output_file)
    output_ws = output_wb.active

    for col_idx in [3, 4, 5, 6]:
        cell = output_ws.cell(1, col_idx)
        if cell.value and isinstance(cell.value, str):
            if any('\u4e00' <= c <= '\u9fff' for c in cell.value):
                try:
                    cell.value = translator.translate(cell.value.strip())
                except Exception as e:
                    print(f"表头翻译错误: {e}")

    output_wb.save(output_file)
    output_wb.close()

    # 批量翻译C列
    while row_idx <= ws.max_row:
        # 收集需要翻译的文本和对应行号
        texts_to_translate = []
        row_indices = []

        for i in range(batch_size):
            if row_idx + i > ws.max_row:
                break

            row = list(ws.iter_rows(min_row=row_idx + i, max_row=row_idx + i))[0]
            c_cell = row[2]

            if c_cell.value and isinstance(c_cell.value, str) and c_cell.value.strip():
                if any('\u4e00' <= c <= '\u9fff' for c in c_cell.value):
                    texts_to_translate.append(c_cell.value.strip())
                    row_indices.append(row_idx + i)

        # 批量翻译
        if texts_to_translate:
            results = translate_batch(translator, texts_to_translate)

            # 写入结果
            if results:
                output_wb = openpyxl.load_workbook(output_file)
                output_ws = output_wb.active

                for idx, row_num in enumerate(row_indices):
                    if idx < len(results):
                        output_ws.cell(row_num, 3).value = results[idx]
                        count += 1

                output_wb.save(output_file)
                output_wb.close()

                print(f"已翻译 {row_idx}-{row_idx + len(row_indices) -1} 行 (共{count}条)")

        row_idx += batch_size
        time.sleep(0.5)  # 批次之间稍作延迟

    wb.close()
    print(f"完成! 共翻译 {count} 行，保存到 {output_file}")


if __name__ == "__main__":
    # 百度翻译API凭证
    appid = '20260325002580664'
    appkey = 'LSEHwM64k2IGRMgX6NaG'

    if len(sys.argv) < 3:
        print("用法: python translate_excel.py <输入文件> <输出文件> [批次大小]")
        print("示例: python translate_excel.py 物质名录.xlsx 物质名录_en.xlsx 100")
    else:
        input_file = sys.argv[1]
        output_file = sys.argv[2]
        batch_size = int(sys.argv[3]) if len(sys.argv) > 3 else 100

        translate_excel(input_file, output_file, appid, appkey, batch_size)